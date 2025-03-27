import streamlit as st
from utils.session import init_session_state  # セッション管理関数のインポート
import pdfplumber
from pdf2image import convert_from_bytes
import pandas as pd
import numpy as np
from PIL import Image
import tempfile
import os
import re
from datetime import datetime, date, timedelta
from openpyxl.utils import get_column_letter
import uuid
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, String, DateTime, Enum, JSON, ForeignKey, Text
from sqlalchemy.sql import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io
import json
import sqlite3
from dataclasses import dataclass
from typing import Optional, Dict, Set, Any
from openpyxl.cell.cell import MergedCell
import hashlib
import base64
import requests
from dotenv import load_dotenv
import secrets
import hmac
import asyncio
import aiohttp
from concurrent.futures import ThreadPoolExecutor
from functools import partial
import logging
import traceback
import sys

# 環境変数の読み込み
load_dotenv()

# Popplerのパス設定
if os.path.exists('/usr/local/bin/pdftoppm'):
    os.environ['PATH'] = f"/usr/local/bin:{os.environ['PATH']}"

# ページ設定（必ず最初に実行）
st.set_page_config(
    page_title="PDF to Excel 変換ツール",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# セッション状態の初期化
init_session_state()  # セッション状態を初期化

if 'rerun_count' not in st.session_state:
    st.session_state.rerun_count = 0
if 'last_rerun_time' not in st.session_state:
    st.session_state.last_rerun_time = datetime.now()
if 'conversion_success' not in st.session_state:
    st.session_state.conversion_success = False
if 'processing_pdf' not in st.session_state:
    st.session_state.processing_pdf = False

# データベースの設定
DB_PATH = "pdf_converter.db"

# プラン定義
PLAN_LIMITS = {
    "free_guest": 3,        # 未ログインユーザー
    "free_registered": 5,   # 登録済み無料ユーザー
    "premium_basic": 1000,  # $5プラン
    "premium_pro": float('inf')  # $20プラン
}

PLAN_NAMES = {
    "free_guest": "無料プラン（未登録）",
    "free_registered": "無料プラン（登録済）",
    "premium_basic": "ベーシックプラン（$5）",
    "premium_pro": "プロフェッショナルプラン（$20）"
}

# ロギングの設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class ErrorHandler:
    """エラーハンドリング用のクラス"""
    
    @staticmethod
    def log_error(error: Exception, context: Dict[str, Any] = None) -> None:
        """エラーをログに記録"""
        error_info = {
            'error_type': type(error).__name__,
            'error_message': str(error),
            'stack_trace': traceback.format_exc(),
            'timestamp': datetime.now().isoformat(),
            'context': context or {}
        }
        
        logger.error(f"エラーが発生しました: {error_info}")
        
        # データベースにも記録
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            
            c.execute('''
                CREATE TABLE IF NOT EXISTS error_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    error_type TEXT,
                    error_message TEXT,
                    stack_trace TEXT,
                    context JSON,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    user_id TEXT,
                    ip_address TEXT,
                    user_agent TEXT
                )
            ''')
            
            c.execute('''
                INSERT INTO error_logs 
                (error_type, error_message, stack_trace, context, user_id, ip_address, user_agent)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                error_info['error_type'],
                error_info['error_message'],
                error_info['stack_trace'],
                json.dumps(error_info['context']),
                st.session_state.get('user_id'),
                st.session_state.get('client_ip'),
                st.session_state.get('user_agent')
            ))
            
            conn.commit()
        except Exception as e:
            logger.error(f"エラーログの保存中にエラーが発生: {str(e)}")
        finally:
            conn.close()
    
    @staticmethod
    def get_user_friendly_message(error: Exception) -> str:
        """ユーザーフレンドリーなエラーメッセージを生成"""
        error_messages = {
            'ValueError': {
                'file_size': 'ファイルサイズが大きすぎます。200MB以下のファイルをアップロードしてください。',
                'pdf_empty': 'PDFファイルが空です。有効なPDFファイルをアップロードしてください。',
                'text_extraction': 'PDFからテキストを抽出できませんでした。スキャンされたPDFの場合は、OCR処理を試みます。',
                'ocr_failed': 'OCR処理に失敗しました。画像の品質を確認してください。',
                'api_key_missing': 'OCRサービスの設定が見つかりません。管理者に連絡してください。'
            },
            'FileNotFoundError': '必要なファイルが見つかりません。管理者に連絡してください。',
            'PermissionError': 'ファイルへのアクセス権限がありません。',
            'sqlite3.Error': 'データベースの操作中にエラーが発生しました。しばらく時間をおいて再度お試しください。',
            'ConnectionError': 'サーバーへの接続に失敗しました。インターネット接続を確認してください。',
            'TimeoutError': '処理がタイムアウトしました。ファイルサイズを小さくして再度お試しください。'
        }
        
        error_type = type(error).__name__
        error_message = str(error)
        
        # エラーメッセージの特定
        if error_type in error_messages:
            if isinstance(error_messages[error_type], dict):
                for key, message in error_messages[error_type].items():
                    if key in error_message.lower():
                        return message
            return error_messages[error_type]
        
        return f"予期せぬエラーが発生しました: {error_message}"

def handle_error(error: Exception, context: Dict[str, Any] = None) -> None:
    """エラーを処理して表示"""
    # エラーのログ記録
    ErrorHandler.log_error(error, context)
    
    # ユーザーフレンドリーなエラーメッセージを表示
    user_message = ErrorHandler.get_user_friendly_message(error)
    st.error(user_message)
    
    # プレミアムユーザーの場合は詳細情報を表示
    if st.session_state.get('is_premium', False):
        with st.expander("エラーの詳細"):
            st.code(traceback.format_exc())

def process_pdf_with_error_handling(uploaded_file, document_type: str, document_date: date = None) -> tuple:
    """エラーハンドリング付きのPDF処理"""
    try:
        return process_pdf_async(uploaded_file, document_type, document_date)
    except Exception as e:
        handle_error(e, {
            'file_name': uploaded_file.name,
            'document_type': document_type,
            'document_date': document_date,
            'file_size': len(uploaded_file.getvalue()) / (1024 * 1024)
        })
        return None, None

class ConversionTracker:
    def __init__(self):
        self.conn = sqlite3.connect('conversion_tracker.db')
        self.setup_database()
    
    def setup_database(self):
        """データベースのセットアップ"""
        with self.conn:
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS conversion_counts (
                    user_id TEXT,
                    ip_address TEXT,
                    browser_id TEXT,
                    conversion_date DATE,
                    count INTEGER,
                    PRIMARY KEY (user_id, ip_address, browser_id, conversion_date)
                )
            """)
            
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS user_plans (
                    user_id TEXT PRIMARY KEY,
                    plan_type TEXT,
                    updated_at TIMESTAMP
                )
            """)
    
    def get_unique_identifier(self, user_id: Optional[str] = None) -> str:
        """ユーザー識別子の生成（IPアドレス + ブラウザID + ユーザーID）"""
        ip = st.session_state.get('client_ip', 'unknown')
        browser_id = st.session_state.get('browser_id')
        if not browser_id:
            browser_id = str(uuid.uuid4())
            st.session_state['browser_id'] = browser_id
        
        identifier = f"{ip}:{browser_id}"
        if user_id:
            identifier += f":{user_id}"
        
        return hashlib.sha256(identifier.encode()).hexdigest()
    
    def get_daily_count(self, user_id: Optional[str] = None) -> int:
        """日次変換回数の取得"""
        identifier = self.get_unique_identifier(user_id)
        today = date.today()
        
        with self.conn:
            cursor = self.conn.execute("""
                SELECT SUM(count) FROM conversion_counts
                WHERE (user_id = ? OR ip_address = ? OR browser_id = ?)
                AND conversion_date = ?
            """, (identifier, identifier, identifier, today))
            
            count = cursor.fetchone()[0]
            return count if count is not None else 0
    
    def increment_count(self, user_id: Optional[str] = None) -> bool:
        """変換回数のインクリメント"""
        identifier = self.get_unique_identifier(user_id)
        today = date.today()
        current_count = self.get_daily_count(user_id)
        plan_limit = self.get_plan_limit(user_id)
        
        if current_count >= plan_limit:
            return False
        
        with self.conn:
            self.conn.execute("""
                INSERT INTO conversion_counts (user_id, ip_address, browser_id, conversion_date, count)
                VALUES (?, ?, ?, ?, 1)
                ON CONFLICT (user_id, ip_address, browser_id, conversion_date)
                DO UPDATE SET count = count + 1
            """, (identifier, identifier, identifier, today))
        
        return True
    
    def get_plan_limit(self, user_id: Optional[str] = None) -> int:
        """ユーザープランの制限値を取得"""
        if not user_id:
            return PLAN_LIMITS["free_guest"]
        
        with self.conn:
            cursor = self.conn.execute("""
                SELECT plan_type FROM user_plans WHERE user_id = ?
            """, (user_id,))
            result = cursor.fetchone()
            
            if result:
                return PLAN_LIMITS.get(result[0], PLAN_LIMITS["free_guest"])
            return PLAN_LIMITS["free_registered"]
    
    def update_plan(self, user_id: str, plan_type: str):
        """ユーザープランの更新"""
        with self.conn:
            self.conn.execute("""
                INSERT INTO user_plans (user_id, plan_type, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT (user_id) DO UPDATE SET
                    plan_type = excluded.plan_type,
                    updated_at = CURRENT_TIMESTAMP
            """, (user_id, plan_type))
    
    def adjust_count_after_registration(self, user_id: str):
        """登録後の変換回数調整（+2回）"""
        identifier = self.get_unique_identifier(user_id)
        today = date.today()
        
        with self.conn:
            # 既存の回数を取得
            cursor = self.conn.execute("""
                SELECT count FROM conversion_counts
                WHERE (user_id = ? OR ip_address = ? OR browser_id = ?)
                AND conversion_date = ?
            """, (identifier, identifier, identifier, today))
            
            current_count = cursor.fetchone()
            if current_count:
                # 既存レコードの更新（最大5回まで）
                new_count = min(current_count[0] + 2, PLAN_LIMITS["free_registered"])
                self.conn.execute("""
                    UPDATE conversion_counts
                    SET count = ?
                    WHERE (user_id = ? OR ip_address = ? OR browser_id = ?)
                    AND conversion_date = ?
                """, (new_count, identifier, identifier, identifier, today))

# グローバルなトラッカーインスタンス
tracker = ConversionTracker()

@dataclass
class ConversionHistory:
    id: int
    user_id: str
    document_type: str
    document_date: str
    conversion_date: datetime
    file_name: str
    status: str

def hash_password(password: str) -> str:
    """パスワードをハッシュ化"""
    salt = os.urandom(32)
    key = hashlib.pbkdf2_hmac(
        'sha256',
        password.encode('utf-8'),
        salt,
        100000
    )
    return base64.b64encode(salt + key).decode('utf-8')

def verify_password(stored_password: str, input_password: str) -> bool:
    """パスワードを検証"""
    try:
        decoded = base64.b64decode(stored_password)
        salt = decoded[:32]
        stored_key = decoded[32:]
        key = hashlib.pbkdf2_hmac(
            'sha256',
            input_password.encode('utf-8'),
            salt,
            100000
        )
        return hmac.compare_digest(stored_key, key)
    except Exception:
        return False

def init_db():
    """データベースの初期化"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # ユーザーテーブル
        c.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id TEXT PRIMARY KEY,
                username TEXT UNIQUE,
                password_hash TEXT,
                email TEXT UNIQUE,
                plan_type TEXT DEFAULT 'free_registered',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP,
                is_active BOOLEAN DEFAULT 1,
                failed_login_attempts INTEGER DEFAULT 0,
                locked_until TIMESTAMP
            )
        ''')
        
        # セッショントークンテーブル
        c.execute('''
            CREATE TABLE IF NOT EXISTS sessions (
                token TEXT PRIMARY KEY,
                user_id TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # 変換履歴テーブル
        c.execute('''
            CREATE TABLE IF NOT EXISTS conversion_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                document_type TEXT,
                document_date TEXT,
                conversion_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                file_name TEXT,
                status TEXT,
                ip_address TEXT,
                user_agent TEXT,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # フィードバックテーブル
        c.execute('''
            CREATE TABLE IF NOT EXISTS feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                rating INTEGER,
                comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status TEXT DEFAULT 'pending',
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        conn.commit()
    except Exception as e:
        st.error(f"データベースの初期化中にエラーが発生しました: {str(e)}")
    finally:
        conn.close()

def create_session(user_id: str) -> str:
    """セッショントークンを生成"""
    token = secrets.token_urlsafe(32)
    expires_at = datetime.now() + timedelta(days=30)
    
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        INSERT INTO sessions (token, user_id, expires_at)
        VALUES (?, ?, ?)
    ''', (token, user_id, expires_at))
    conn.commit()
    conn.close()
    
    return token

def verify_session(token: str) -> Optional[str]:
    """セッショントークンを検証"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT user_id FROM sessions
        WHERE token = ? AND expires_at > CURRENT_TIMESTAMP
    ''', (token,))
    result = c.fetchone()
    conn.close()
    
    return result[0] if result else None

def handle_login(username: str, password: str) -> bool:
    """ログイン処理を実行"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # ユーザー情報の取得
        c.execute('''
            SELECT id, password_hash, is_active, failed_login_attempts, locked_until
            FROM users WHERE username = ?
        ''', (username,))
        result = c.fetchone()
        
        if not result:
            st.error("ユーザー名またはパスワードが正しくありません。")
            return False
            
        user_id, stored_hash, is_active, failed_attempts, locked_until = result
        
        # アカウントロックのチェック
        if not is_active:
            st.error("このアカウントは無効化されています。")
            return False
            
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now():
            st.error("アカウントが一時的にロックされています。しばらく時間をおいて再度お試しください。")
            return False
            
        # パスワードの検証
        if not verify_password(stored_hash, password):
            # ログイン失敗回数を更新
            c.execute('''
                UPDATE users
                SET failed_login_attempts = failed_login_attempts + 1,
                    locked_until = CASE
                        WHEN failed_login_attempts >= 4 THEN datetime('now', '+15 minutes')
                        ELSE locked_until
                    END
                WHERE id = ?
            ''', (user_id,))
            conn.commit()
            st.error("ユーザー名またはパスワードが正しくありません。")
            return False
            
        # ログイン成功時の処理
        c.execute('''
            UPDATE users
            SET last_login = CURRENT_TIMESTAMP,
                failed_login_attempts = 0,
                locked_until = NULL
            WHERE id = ?
        ''', (user_id,))
        conn.commit()
        
        # セッショントークンの生成
        token = create_session(user_id)
        
        # セッション状態の更新
        st.session_state.logged_in = True
        st.session_state.user_id = user_id
        st.session_state.username = username
        st.session_state.session_token = token
        
        return True
        
    except Exception as e:
        st.error(f"ログイン処理中にエラーが発生しました: {str(e)}")
        return False
    finally:
        conn.close()

def handle_logout():
    """ログアウト処理を実行"""
    try:
        if 'session_token' in st.session_state:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute('DELETE FROM sessions WHERE token = ?', (st.session_state.session_token,))
            conn.commit()
            conn.close()
        
        # セッション状態のクリア
        for key in ['logged_in', 'user_id', 'username', 'session_token']:
            if key in st.session_state:
                del st.session_state[key]
        
        st.success("ログアウトしました。")
        st.rerun()
        
    except Exception as e:
        st.error(f"ログアウト処理中にエラーが発生しました: {str(e)}")

def handle_register(username: str, password: str, email: str) -> bool:
    """新規登録処理を実行"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # ユーザー名とメールアドレスの重複チェック
        c.execute('SELECT id FROM users WHERE username = ? OR email = ?', (username, email))
        if c.fetchone():
            st.error("このユーザー名またはメールアドレスは既に使用されています。")
            return False
        
        # パスワードのハッシュ化
        password_hash = hash_password(password)
        
        # ユーザーの作成
        user_id = str(uuid.uuid4())
        c.execute('''
            INSERT INTO users (id, username, password_hash, email, plan_type)
            VALUES (?, ?, ?, ?, 'free_registered')
        ''', (user_id, username, password_hash, email))
        conn.commit()
        
        # セッショントークンの生成
        token = create_session(user_id)
        
        # セッション状態の更新
        st.session_state.logged_in = True
        st.session_state.user_id = user_id
        st.session_state.username = username
        st.session_state.session_token = token
        
        return True
        
    except Exception as e:
        st.error(f"新規登録処理中にエラーが発生しました: {str(e)}")
        return False
    finally:
        conn.close()

def save_conversion_history(user_id: str, document_type: str, document_date: str, file_name: str, status: str):
    """変換履歴を保存"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        INSERT INTO conversion_history 
        (user_id, document_type, document_date, file_name, status)
        VALUES (?, ?, ?, ?, ?)
    ''', (user_id, document_type, document_date, file_name, status))
    conn.commit()
    conn.close()

def get_daily_conversion_count(user_id: str) -> int:
    """ユーザーの本日の変換回数を取得"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        today = datetime.now().strftime('%Y-%m-%d')
        
        # ユーザーと日付の組み合わせがなければ作成
        c.execute('''
            INSERT OR IGNORE INTO conversion_count (user_id, count_date, count)
            VALUES (?, ?, 0)
        ''', (user_id, today))
        
        # カウントを取得
        c.execute('''
            SELECT count FROM conversion_count
            WHERE user_id = ? AND count_date = ?
        ''', (user_id, today))
        
        result = c.fetchone()
        conn.commit()
        return result[0] if result else 0
    except Exception as e:
        st.error(f"変換回数の取得中にエラーが発生しました: {str(e)}")
        return 0
    finally:
        conn.close()

def check_and_increment_conversion_count(user_id: Optional[str] = None) -> bool:
    """変換回数をチェックしてインクリメント"""
    try:
        # 現在の変換回数と制限を取得
        daily_count = tracker.get_daily_count(user_id)
        limit = tracker.get_plan_limit(user_id)
        
        # 制限チェック
        if daily_count >= limit:
            user_plan = st.session_state.get('user_plan', 'free_guest')
            if user_plan == 'free_guest':
                st.error("""
                本日の変換回数制限に達しました。
                アカウントを作成すると、1日5回まで変換可能になります。
                """)
            elif user_plan == 'free_registered':
                st.error("""
                本日の変換回数制限に達しました。
                プレミアムプランにアップグレードすると、無制限に変換可能になります。
                """)
            else:
                st.error("本日の変換回数制限に達しました。")
            return False
            
        # カウントをインクリメント
        if tracker.increment_count(user_id):
            st.session_state.conversion_success = True
            st.success("変換が完了しました！")
            
            # 画面の更新（最大1回まで）
            current_time = datetime.now()
            if (current_time - st.session_state.last_rerun_time).total_seconds() > 1:
                st.session_state.rerun_count += 1
                if st.session_state.rerun_count <= 1:
                    st.session_state.last_rerun_time = current_time
                    st.rerun()
            return True
            
        st.error("変換回数の更新に失敗しました。しばらく時間をおいて再度お試しください。")
        return False
        
    except Exception as e:
        st.error(f"""
        変換回数の処理中にエラーが発生しました。
        エラー内容: {str(e)}
        しばらく時間をおいて再度お試しください。
        """)
        return False

def get_user_plan(user_id):
    """ユーザーのプランを取得する関数"""
    try:
        if user_id is None:
            return "free_guest"
        
        # セッションからプラン情報を取得
        user_plan = st.session_state.get('user_plan', 'free_registered')
        
        # プレミアムユーザーの判定（仮の実装）
        premium_users = st.session_state.get('premium_users', set())
        if user_id in premium_users:
            return "premium"
        
        return user_plan
    except Exception as e:
        st.error(f"プラン情報の取得中にエラーが発生しました: {str(e)}")
        return "free_guest"  # エラー時は最も制限の厳しいプランを返す

def get_plan_limits(plan_type):
    """プランごとの制限を取得"""
    limits = {
        "premium": float('inf'),  # 無制限
        "free_registered": 5,     # ログインユーザー
        "free_guest": 3          # 未ログインユーザー
    }
    return limits.get(plan_type, 3)  # デフォルトは3回

def get_conversion_limit(user_id=None):
    """ユーザーの変換制限を取得"""
    plan = get_user_plan(user_id)
    return get_plan_limits(plan)

@st.cache_data(ttl=300)
def optimize_image(image_bytes: bytes, max_size: int = 1024) -> bytes:
    """画像を最適化（5分キャッシュ）"""
    try:
        # 画像を開く
        img = Image.open(io.BytesIO(image_bytes))
        
        # 画像サイズを取得
        width, height = img.size
        
        # 最大サイズを超えている場合、リサイズ
        if width > max_size or height > max_size:
            ratio = min(max_size / width, max_size / height)
            new_size = (int(width * ratio), int(height * ratio))
            img = img.resize(new_size, Image.Resampling.LANCZOS)
        
        # 画像を最適化して保存
        output = io.BytesIO()
        img.save(output, format='PNG', optimize=True, quality=85)
        return output.getvalue()
    except Exception as e:
        st.error(f"画像の最適化中にエラーが発生しました: {str(e)}")
        return image_bytes

@st.cache_data(ttl=300)
def create_preview(pdf_file):
    """PDFプレビューを生成"""
    try:
        # PDFを画像に変換
        images = convert_from_bytes(pdf_file.read())
        if not images:
            raise ValueError("PDFファイルが空です")
        
        # 最初のページのみを表示
        first_page = images[0]
        
        # 画像を最適化
        img_byte_arr = io.BytesIO()
        first_page.save(img_byte_arr, format='PNG')
        img_bytes = optimize_image(img_byte_arr.getvalue())
        
        return io.BytesIO(img_bytes)
        
    except Exception as e:
        if "poppler" in str(e).lower():
            st.error("""
            PDFプレビューの生成に失敗しました。
            Popplerがインストールされていないか、PATHが通っていません。
            
            インストール方法：
            ```bash
            brew install poppler
            ```
            """)
        else:
            st.error(f"プレビューの生成中にエラーが発生しました: {str(e)}")
        return None

@st.cache_data(ttl=3600)
def get_document_type_label(doc_type: str) -> str:
    """ドキュメントタイプのラベルを取得（1時間キャッシュ）"""
    type_map = {
        "estimate": "見積書",
        "invoice": "請求書",
        "delivery": "納品書",
        "receipt": "領収書",
        "financial": "決算書",
        "tax_return": "確定申告書",
        "other": "その他"
    }
    return type_map.get(doc_type, "不明な書類")

@st.cache_data(ttl=300)
def process_pdf_with_ocr(image_bytes: bytes, document_type: str) -> str:
    """OCR処理を実行（5分キャッシュ）"""
    try:
        # 画像を最適化
        optimized_image = optimize_image(image_bytes)
        
        # 画像をbase64エンコード
        image_base64 = base64.b64encode(optimized_image).decode('utf-8')
        
        # APIキーの取得
        api_key = os.getenv('GOOGLE_VISION_API_KEY')
        if not api_key:
            st.error("Google Cloud Vision APIの設定が見つかりません。")
            return None
            
        # APIエンドポイントの設定
        url = f"https://vision.googleapis.com/v1/images:annotate?key={api_key}"
        headers = {"Content-Type": "application/json"}
        
        # リクエストデータの作成
        data = {
            "requests": [{
                "image": {"content": image_base64},
                "features": [{"type": "DOCUMENT_TEXT_DETECTION"}]
            }]
        }
        
        # APIリクエストの送信
        response = requests.post(url, headers=headers, json=data)
        
        # レスポンスの確認
        if response.status_code != 200:
            st.error(f"OCR処理に失敗しました。")
            return None
            
        # レスポンスの解析
        result = response.json()
        if 'responses' in result and result['responses']:
            text_annotations = result['responses'][0].get('textAnnotations', [])
            if text_annotations:
                return text_annotations[0].get('description', '')
                
        return None
        
    except Exception as e:
        st.error(f"OCR処理中にエラーが発生しました: {str(e)}")
        return None

@st.cache_data(ttl=3600)
def create_excel_file(text_content: str, document_type: str, document_date: date = None) -> tuple:
    """Excelファイルを作成（5分キャッシュ）"""
    try:
        # ドキュメントタイプの日本語名を取得
        doc_type_ja = get_document_type_label(document_type)
        
        # 日付の処理
        if document_date:
            date_str = document_date.strftime("%Y%m%d")
        else:
            date_str = datetime.now().strftime("%Y%m%d")
        
        # Excelワークブックの作成
        wb = Workbook()
        ws = wb.active
        ws.title = f"{doc_type_ja}_{date_str}"
        
        # フォントとスタイルの設定
        default_font = Font(name='Yu Gothic', size=11)
        header_font = Font(name='Yu Gothic', size=12, bold=True)
        title_font = Font(name='Yu Gothic', size=14, bold=True)
        
        # 基本的なセルスタイル
        normal_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 罫線スタイル
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # テキストを行に分割
        lines = text_content.split('\n')
        
        # 行の高さと列幅の初期設定
        ws.row_dimensions[1].height = 30
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # データの書き込みと書式設定
        for i, line in enumerate(lines, 1):
            ws.cell(row=i, column=1, value=line)
            cell = ws[f"A{i}"]
            cell.font = default_font
            cell.alignment = normal_alignment
            cell.border = thin_border
            
            # 金額と思われる部分は右寄せに
            if any(char in line for char in ['¥', '円', '税']):
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # タイトル行の特別な書式設定
        if len(lines) > 0:
            title_cell = ws['A1']
            title_cell.font = title_font
            title_cell.alignment = center_alignment
            title_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # メモリ上にExcelファイルを保存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # ファイル名の生成
        file_name = f"{doc_type_ja}_{date_str}.xlsx"
        
        return excel_buffer, file_name
    except Exception as e:
        st.error(f"Excelファイルの作成中にエラーが発生しました: {str(e)}")
        return None, None

def display_conversion_count():
    """変換回数を表示"""
    if 'conversion_count' not in st.session_state:
        st.session_state.conversion_count = 0
    
    if st.session_state.logged_in:
        if st.session_state.get('is_premium', False):
            st.info("本日の変換回数：無制限（有料プラン）")
        else:
            st.info(f"本日の変換回数：{st.session_state.conversion_count} / 5回（無料プラン・登録済）")
    else:
        st.info(f"本日の変換回数：{st.session_state.conversion_count} / 3回（未登録）")

def create_document_type_buttons():
    """ドキュメントタイプ選択ボタンを作成"""
    st.write("ドキュメントの種類を選択")
    
    # セッション状態の初期化
    if 'selected_document_type' not in st.session_state:
        st.session_state.selected_document_type = None
    
    # ラジオボタンをボタン風にするスタイル
    button_style = """
        <style>
        /* ラジオボタンを非表示 */
        div[data-testid="stRadio"] > div > div > label > div:first-child {
            display: none;
        }
        
        /* ラベルをボタン風にスタイリング */
        div[data-testid="stRadio"] label {
            width: 100%;
            min-height: 60px;
            margin: 8px 0;
            padding: 10px 15px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            background: #ffffff;
            font-size: 16px;
            color: #333;
            transition: all 0.2s ease;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            display: flex;
            align-items: center;
            justify-content: center;
            cursor: pointer;
            user-select: none;
        }
        
        /* ホバー時のスタイル */
        div[data-testid="stRadio"] label:hover {
            border-color: #2196F3;
            box-shadow: 0 4px 8px rgba(33,150,243,0.2);
        }
        
        /* 選択時のスタイル */
        div[data-testid="stRadio"] label[data-checked="true"] {
            border-color: #2196F3;
            background: #e3f2fd;
            color: #1565C0;
            box-shadow: 0 4px 8px rgba(33,150,243,0.2);
        }
        
        /* 2カラムレイアウト */
        div[data-testid="stRadio"] > div {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 10px;
        }

        /* 変換ボタンのスタイル */
        .stButton > button {
            width: 100%;
            padding: 15px 30px;
            background: linear-gradient(145deg, #2196F3 0%, #1976D2 100%);
            color: white;
            border: none;
            border-radius: 25px;
            font-size: 18px;
            font-weight: bold;
            transition: all 0.3s ease;
            box-shadow: 0 4px 6px rgba(33,150,243,0.2);
            margin-top: 20px;
        }

        .stButton > button:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 12px rgba(33,150,243,0.3);
            background: linear-gradient(145deg, #1976D2 0%, #1565C0 100%);
        }

        /* エラーメッセージのスタイル */
        .stAlert {
            background: #ffebee;
            color: #c62828;
            padding: 1rem;
            border-radius: 8px;
            border-left: 4px solid #c62828;
            margin: 1rem 0;
        }

        /* 成功メッセージのスタイル */
        .stSuccess {
            background: #e8f5e9;
            color: #2e7d32;
            padding: 1rem;
            border-radius: 8px;
            border-left: 4px solid #2e7d32;
            margin: 1rem 0;
        }
        </style>
    """
    st.markdown(button_style, unsafe_allow_html=True)
    
    # ドキュメントタイプの定義
    document_types = {
        "見積書": "estimate",
        "請求書": "invoice",
        "納品書": "delivery",
        "領収書": "receipt",
        "決算書": "financial",
        "確定申告書": "tax_return",
        "その他": "other"
    }
    
    # ラジオボタンで選択
    selected_label = st.radio(
        "書類の種類",
        options=list(document_types.keys()),
        key="doc_type_radio",
        label_visibility="collapsed",  # ラベルを非表示
        horizontal=True,  # 水平配置
        index=None if st.session_state.selected_document_type is None else 
              list(document_types.values()).index(st.session_state.selected_document_type)
    )
    
    # 選択状態の更新
    if selected_label is not None:
        st.session_state.selected_document_type = document_types[selected_label]
    else:
        st.warning("書類の種類を選択してください")
        return None
    
    return st.session_state.selected_document_type

def create_footer():
    """フッターを作成"""
    st.markdown("---")
    
    # 2カラムレイアウトでフッターを表示
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("### 📄 関連リンク")
        st.markdown("""
        - [利用規約](/terms) - サービスの利用条件について
        - [プライバシーポリシー](/privacy) - 個人情報の取り扱いについて
        - [お問い合わせ](/contact) - ご質問・ご要望はこちら
        """)
    
    with col2:
        st.markdown("### 💡 プラン情報")
        if st.session_state.logged_in:
            if st.session_state.get('is_premium', False):
                st.markdown("**プレミアムプラン**")
                st.markdown("- 無制限の変換")
                st.markdown("- 高精度OCR")
                st.markdown("- 広告非表示")
            else:
                st.markdown("**無料プラン**")
                st.markdown("- 1日5回まで")
                st.markdown("- 基本OCR")
        else:
            st.markdown("**無料プラン**")
            st.markdown("- 1日3回まで")
            st.markdown("- 基本OCR")

def create_hero_section():
    """ヒーローセクションを作成"""
    st.markdown("""
        <style>
        .hero-section {
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            padding: 4rem 2rem;
            border-radius: 1rem;
            margin-bottom: 2rem;
            color: white;
            text-align: center;
        }
        .hero-title {
            font-size: 2.5rem;
            font-weight: bold;
            margin-bottom: 1rem;
        }
        .hero-description {
            font-size: 1.2rem;
            margin-bottom: 2rem;
            opacity: 0.9;
        }
        .feature-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 1.5rem;
            margin-top: 2rem;
        }
        .feature-card {
            background: rgba(255, 255, 255, 0.1);
            padding: 1.5rem;
            border-radius: 0.5rem;
            transition: transform 0.3s ease;
        }
        .feature-card:hover {
            transform: translateY(-5px);
        }
        .feature-icon {
            font-size: 2rem;
            margin-bottom: 1rem;
        }
        .feature-title {
            font-size: 1.2rem;
            font-weight: bold;
            margin-bottom: 0.5rem;
        }
        .feature-description {
            font-size: 0.9rem;
            opacity: 0.8;
        }
        @media (max-width: 768px) {
            .hero-section {
                padding: 2rem 1rem;
            }
            .hero-title {
                font-size: 2rem;
            }
            .hero-description {
                font-size: 1rem;
            }
            .feature-grid {
                grid-template-columns: 1fr;
            }
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="hero-section">
            <h1 class="hero-title">PDF to Excel 変換ツール</h1>
            <p class="hero-description">
                スキャンされたPDFや画像からテキストを抽出し、Excelファイルに変換します。<br>
                請求書、領収書、報告書など、様々なPDFに対応。
            </p>
            <div class="feature-grid">
                <div class="feature-card">
                    <div class="feature-icon">📄</div>
                    <h3 class="feature-title">様々なPDFに対応</h3>
                    <p class="feature-description">スキャンされたPDF、画像PDF、テキストPDFなど、幅広い形式に対応</p>
                </div>
                <div class="feature-card">
                    <div class="feature-icon">🔍</div>
                    <h3 class="feature-title">高精度OCR</h3>
                    <p class="feature-description">Google Cloud Vision APIを使用した高精度なテキスト認識</p>
                </div>
                <div class="feature-card">
                    <div class="feature-icon">⚡</div>
                    <h3 class="feature-title">高速処理</h3>
                    <p class="feature-description">非同期処理による高速なファイル変換</p>
                </div>
                <div class="feature-card">
                    <div class="feature-icon">📊</div>
                    <h3 class="feature-title">データ分析</h3>
                    <p class="feature-description">変換履歴の分析とレポート生成機能</p>
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)

def create_login_section():
    """ログインセクションを作成"""
    st.markdown("""
        <style>
        .login-section {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 2rem;
            margin: 2rem 0;
        }
        .login-form, .register-form {
            background: white;
            padding: 2rem;
            border-radius: 0.5rem;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .form-title {
            font-size: 1.5rem;
            font-weight: bold;
            margin-bottom: 1.5rem;
            color: #1e3c72;
        }
        .form-group {
            margin-bottom: 1rem;
        }
        .form-label {
            display: block;
            margin-bottom: 0.5rem;
            color: #666;
        }
        .form-input {
            width: 100%;
            padding: 0.5rem;
            border: 1px solid #ddd;
            border-radius: 0.25rem;
            font-size: 1rem;
        }
        .form-input:focus {
            outline: none;
            border-color: #1e3c72;
            box-shadow: 0 0 0 2px rgba(30,60,114,0.1);
        }
        .form-button {
            width: 100%;
            padding: 0.75rem;
            background: #1e3c72;
            color: white;
            border: none;
            border-radius: 0.25rem;
            font-size: 1rem;
            cursor: pointer;
            transition: background 0.3s ease;
        }
        .form-button:hover {
            background: #2a5298;
        }
        .form-link {
            color: #1e3c72;
            text-decoration: none;
            font-size: 0.9rem;
        }
        .form-link:hover {
            text-decoration: underline;
        }
        @media (max-width: 768px) {
            .login-section {
                grid-template-columns: 1fr;
            }
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="login-section">
            <div class="login-form">
                <h2 class="form-title">ログイン</h2>
                <form>
                    <div class="form-group">
                        <label class="form-label" for="login-email">メールアドレス</label>
                        <input type="email" id="login-email" class="form-input" required>
                    </div>
                    <div class="form-group">
                        <label class="form-label" for="login-password">パスワード</label>
                        <input type="password" id="login-password" class="form-input" required>
                    </div>
                    <button type="submit" class="form-button">ログイン</button>
                </form>
            </div>
            <div class="register-form">
                <h2 class="form-title">新規登録</h2>
                <form>
                    <div class="form-group">
                        <label class="form-label" for="register-email">メールアドレス</label>
                        <input type="email" id="register-email" class="form-input" required>
                    </div>
                    <div class="form-group">
                        <label class="form-label" for="register-password">パスワード</label>
                        <input type="password" id="register-password" class="form-input" required>
                    </div>
                    <div class="form-group">
                        <label class="form-label" for="register-password-confirm">パスワード（確認）</label>
                        <input type="password" id="register-password-confirm" class="form-input" required>
                    </div>
                    <button type="submit" class="form-button">登録</button>
                </form>
            </div>
        </div>
    """, unsafe_allow_html=True)

def create_register_section():
    """新規登録セクションを作成"""
    st.markdown("### 📝 新規アカウント登録")
    with st.form("register_form"):
        new_username = st.text_input("ユーザー名")
        new_password = st.text_input("パスワード", type="password")
        confirm_password = st.text_input("パスワード（確認）", type="password")
        submitted = st.form_submit_button("登録")
        
        if submitted:
            if new_username and new_password and confirm_password:
                if new_password == confirm_password:
                    if handle_register(new_username, new_password, st.session_state.get('email')):
                        st.success("アカウントが作成されました！")
                        st.session_state.logged_in = True
                        st.rerun()
                    else:
                        st.error("アカウントの作成に失敗しました。")
                else:
                    st.error("パスワードが一致しません。")
            else:
                st.error("すべての項目を入力してください。")
    
    if st.button("ログイン画面に戻る"):
        st.session_state.show_register = False
        st.rerun()

def show_ads():
    """広告を表示（プロプランとプレミアムユーザー以外の場合のみ）"""
    user_plan = st.session_state.get('user_plan', 'free_guest')
    if user_plan not in ['premium_basic', 'premium_pro']:
        st.markdown("""
        <script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client=ca-pub-9624397569723291"
             crossorigin="anonymous"></script>
        """, unsafe_allow_html=True)
        
        # ファイルアップロードボタンの下に広告を表示
        st.markdown("""
        <ins class="adsbygoogle"
             style="display:block"
             data-ad-client="ca-pub-9624397569723291"
             data-ad-slot="XXXXXXXXXX"
             data-ad-format="auto"
             data-full-width-responsive="true"></ins>
        """, unsafe_allow_html=True)

def save_feedback(user_id: str, rating: int, comment: str) -> bool:
    """フィードバックを保存"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            INSERT INTO feedback (user_id, rating, comment)
            VALUES (?, ?, ?)
        ''', (user_id, rating, comment))
        conn.commit()
        return True
    except Exception as e:
        st.error(f"フィードバックの保存中にエラーが発生しました: {str(e)}")
        return False
    finally:
        conn.close()

def create_feedback_section():
    """フィードバックセクションを作成"""
    st.markdown("""
        <style>
        .feedback-section {
            background: white;
            padding: 2rem;
            border-radius: 0.5rem;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 2rem 0;
        }
        .feedback-title {
            font-size: 1.5rem;
            font-weight: bold;
            color: #1e3c72;
            margin-bottom: 1.5rem;
            text-align: center;
        }
        .rating-container {
            display: flex;
            justify-content: center;
            gap: 1rem;
            margin-bottom: 1.5rem;
        }
        .rating-button {
            background: none;
            border: none;
            font-size: 2rem;
            cursor: pointer;
            padding: 0.5rem;
            transition: transform 0.3s ease;
        }
        .rating-button:hover {
            transform: scale(1.2);
        }
        .feedback-textarea {
            width: 100%;
            min-height: 100px;
            padding: 0.5rem;
            border: 1px solid #ddd;
            border-radius: 0.25rem;
            margin-bottom: 1rem;
            font-size: 1rem;
        }
        .feedback-textarea:focus {
            outline: none;
            border-color: #1e3c72;
            box-shadow: 0 0 0 2px rgba(30,60,114,0.1);
        }
        .feedback-submit {
            background: #1e3c72;
            color: white;
            border: none;
            padding: 0.75rem 1.5rem;
            border-radius: 0.25rem;
            cursor: pointer;
            font-size: 1rem;
            transition: background 0.3s ease;
        }
        .feedback-submit:hover {
            background: #2a5298;
        }
        @media (max-width: 768px) {
            .feedback-section {
                padding: 1rem;
            }
            .rating-container {
                flex-wrap: wrap;
            }
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="feedback-section">
            <h2 class="feedback-title">フィードバック</h2>
            <p>変換結果はいかがでしたか？評価をお願いします。</p>
            <div class="rating-container">
                <button class="rating-button" onclick="setRating(1)">😢</button>
                <button class="rating-button" onclick="setRating(2)">😕</button>
                <button class="rating-button" onclick="setRating(3)">😐</button>
                <button class="rating-button" onclick="setRating(4)">🙂</button>
                <button class="rating-button" onclick="setRating(5)">😊</button>
            </div>
            <textarea class="feedback-textarea" placeholder="コメントがあれば入力してください"></textarea>
            <button class="feedback-submit" onclick="submitFeedback()">送信</button>
        </div>
    """, unsafe_allow_html=True)

def generate_report(user_id: str, start_date: date, end_date: date) -> tuple:
    """レポートを生成"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # 変換履歴の集計
        c.execute('''
            SELECT 
                document_type,
                COUNT(*) as total_count,
                SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) as success_count,
                SUM(CASE WHEN status LIKE 'error%' THEN 1 ELSE 0 END) as error_count
            FROM conversion_history
            WHERE user_id = ?
            AND date(conversion_date) BETWEEN ? AND ?
            GROUP BY document_type
        ''', (user_id, start_date, end_date))
        conversion_stats = c.fetchall()
        
        # フィードバックの集計
        c.execute('''
            SELECT 
                rating,
                COUNT(*) as count
            FROM feedback
            WHERE user_id = ?
            AND date(created_at) BETWEEN ? AND ?
            GROUP BY rating
        ''', (user_id, start_date, end_date))
        feedback_stats = c.fetchall()
        
        # 日次変換数の推移
        c.execute('''
            SELECT 
                date(conversion_date) as date,
                COUNT(*) as count
            FROM conversion_history
            WHERE user_id = ?
            AND date(conversion_date) BETWEEN ? AND ?
            GROUP BY date(conversion_date)
            ORDER BY date
        ''', (user_id, start_date, end_date))
        daily_conversions = c.fetchall()
        
        return conversion_stats, feedback_stats, daily_conversions
    except Exception as e:
        st.error(f"レポートの生成中にエラーが発生しました: {str(e)}")
        return [], [], []
    finally:
        conn.close()

def create_report_section():
    """レポートセクションを作成"""
    st.markdown("""
        <style>
        .report-section {
            background: white;
            padding: 2rem;
            border-radius: 0.5rem;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin: 2rem 0;
        }
        .report-title {
            font-size: 1.5rem;
            font-weight: bold;
            color: #1e3c72;
            margin-bottom: 1.5rem;
        }
        .report-filters {
            display: flex;
            gap: 1rem;
            margin-bottom: 1.5rem;
            flex-wrap: wrap;
        }
        .report-filter {
            flex: 1;
            min-width: 200px;
        }
        .report-chart {
            margin: 2rem 0;
            padding: 1rem;
            background: #f8f9fa;
            border-radius: 0.5rem;
        }
        .report-table {
            width: 100%;
            border-collapse: collapse;
            margin: 1rem 0;
        }
        .report-table th,
        .report-table td {
            padding: 0.75rem;
            border: 1px solid #dee2e6;
            text-align: left;
        }
        .report-table th {
            background: #f8f9fa;
            font-weight: bold;
        }
        .report-table tr:nth-child(even) {
            background: #f8f9fa;
        }
        .report-download {
            background: #1e3c72;
            color: white;
            border: none;
            padding: 0.75rem 1.5rem;
            border-radius: 0.25rem;
            cursor: pointer;
            font-size: 1rem;
            transition: background 0.3s ease;
        }
        .report-download:hover {
            background: #2a5298;
        }
        @media (max-width: 768px) {
            .report-section {
                padding: 1rem;
            }
            .report-filters {
                flex-direction: column;
            }
            .report-filter {
                width: 100%;
            }
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="report-section">
            <h2 class="report-title">利用レポート</h2>
            <div class="report-filters">
                <div class="report-filter">
                    <label>期間</label>
                    <select>
                        <option>過去7日間</option>
                        <option>過去30日間</option>
                        <option>過去90日間</option>
                    </select>
                </div>
                <div class="report-filter">
                    <label>ドキュメントタイプ</label>
                    <select>
                        <option>すべて</option>
                        <option>請求書</option>
                        <option>領収書</option>
                        <option>報告書</option>
                    </select>
                </div>
            </div>
            <div class="report-chart">
                <h3>変換件数の推移</h3>
                <!-- ここにグラフを表示 -->
            </div>
            <table class="report-table">
                <thead>
                    <tr>
                        <th>日付</th>
                        <th>変換件数</th>
                        <th>成功率</th>
                        <th>平均評価</th>
                    </tr>
                </thead>
                <tbody>
                    <!-- ここにデータを表示 -->
                </tbody>
            </table>
            <button class="report-download">レポートをダウンロード</button>
        </div>
    """, unsafe_allow_html=True)

def create_alert_table():
    """アラートテーブルを作成"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
            CREATE TABLE IF NOT EXISTS alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                alert_type TEXT,
                message TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_read BOOLEAN DEFAULT 0,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        conn.commit()
    except Exception as e:
        st.error(f"アラートテーブルの作成中にエラーが発生しました: {str(e)}")
    finally:
        conn.close()

def create_alert(user_id: str, alert_type: str, message: str) -> bool:
    """アラートを作成"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
            INSERT INTO alerts (user_id, alert_type, message)
            VALUES (?, ?, ?)
        ''', (user_id, alert_type, message))
        
        conn.commit()
        return True
    except Exception as e:
        st.error(f"アラートの作成中にエラーが発生しました: {str(e)}")
        return False
    finally:
        conn.close()

def get_user_alerts(user_id: str) -> list:
    """ユーザーのアラートを取得"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
            SELECT id, alert_type, message, created_at, is_read
            FROM alerts
            WHERE user_id = ?
            ORDER BY created_at DESC
            LIMIT 10
        ''', (user_id,))
        
        alerts = c.fetchall()
        return alerts
    except Exception as e:
        st.error(f"アラートの取得中にエラーが発生しました: {str(e)}")
        return []
    finally:
        conn.close()

def mark_alert_as_read(alert_id: int) -> bool:
    """アラートを既読にする"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
            UPDATE alerts
            SET is_read = 1
            WHERE id = ?
        ''', (alert_id,))
        
        conn.commit()
        return True
    except Exception as e:
        st.error(f"アラートの更新中にエラーが発生しました: {str(e)}")
        return False
    finally:
        conn.close()

def create_alert_section():
    """アラートセクションを作成"""
    st.markdown("""
        <style>
        .alert-section {
            background: #fff3cd;
            border: 1px solid #ffeeba;
            border-radius: 0.5rem;
            padding: 1rem;
            margin-bottom: 1rem;
        }
        .alert-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 0.5rem;
        }
        .alert-title {
            font-weight: bold;
            color: #856404;
        }
        .alert-close {
            background: none;
            border: none;
            color: #856404;
            cursor: pointer;
            font-size: 1.2rem;
        }
        .alert-content {
            color: #856404;
        }
        .alert-list {
            list-style: none;
            padding: 0;
            margin: 0;
        }
        .alert-item {
            padding: 0.5rem 0;
            border-bottom: 1px solid #ffeeba;
        }
        .alert-item:last-child {
            border-bottom: none;
        }
        </style>
    """, unsafe_allow_html=True)
    
    alerts = get_user_alerts(st.session_state.user_id)
    if alerts:
        st.markdown("""
            <div class="alert-section">
                <div class="alert-header">
                    <h3 class="alert-title">お知らせ</h3>
                    <button class="alert-close" onclick="this.parentElement.parentElement.style.display='none'">×</button>
                </div>
                <div class="alert-content">
                    <ul class="alert-list">
        """, unsafe_allow_html=True)
        
        for alert in alerts:
            st.markdown(f"""
                <li class="alert-item">
                    {alert['message']}
                    <small>({alert['created_at']})</small>
                </li>
            """, unsafe_allow_html=True)
        
        st.markdown("""
                    </ul>
                </div>
            </div>
        """, unsafe_allow_html=True)

def check_conversion_limits():
    """変換回数制限のチェックとアラート"""
    if not st.session_state.logged_in:
        return
    
    daily_count = tracker.get_daily_count(st.session_state.user_id)
    limit = tracker.get_plan_limit(st.session_state.user_id)
    
    # 制限に近づいている場合のアラート
    if daily_count >= limit * 0.8:  # 80%以上使用
        create_alert(
            st.session_state.user_id,
            "warning",
            f"本日の変換回数が制限の80%に達しています。残り{limit - daily_count}回です。"
        )
    
    # 制限に達した場合のアラート
    if daily_count >= limit:
        create_alert(
            st.session_state.user_id,
            "error",
            "本日の変換回数制限に達しました。プランをアップグレードすると、より多くの変換が可能です。"
        )

def check_system_status():
    """システムステータスのチェックとアラート"""
    try:
        # OCRサービスの状態チェック
        api_key = os.getenv('GOOGLE_VISION_API_KEY')
        if not api_key:
            create_alert(
                st.session_state.user_id,
                "error",
                "OCRサービスの設定が見つかりません。管理者に連絡してください。"
            )
        
        # データベース接続チェック
        conn = sqlite3.connect(DB_PATH)
        conn.close()
        
    except Exception as e:
        create_alert(
            st.session_state.user_id,
            "error",
            f"システムエラーが発生しました: {str(e)}"
        )

def optimize_database():
    """データベースの最適化を実行"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # インデックスの作成
        c.execute('''
            CREATE INDEX IF NOT EXISTS idx_conversion_history_user_date 
            ON conversion_history(user_id, conversion_date)
        ''')
        
        c.execute('''
            CREATE INDEX IF NOT EXISTS idx_feedback_user_date 
            ON feedback(user_id, created_at)
        ''')
        
        c.execute('''
            CREATE INDEX IF NOT EXISTS idx_alerts_user_read 
            ON alerts(user_id, is_read, created_at)
        ''')
        
        # テーブルの最適化
        c.execute('VACUUM')
        c.execute('ANALYZE')
        
        conn.commit()
    except Exception as e:
        st.error(f"データベースの最適化中にエラーが発生しました: {str(e)}")
    finally:
        conn.close()

@st.cache_data(ttl=3600)
def get_user_stats(user_id: str) -> dict:
    """ユーザーの統計情報を取得（1時間キャッシュ）"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # インデックスを活用した効率的なクエリ
        c.execute('''
            WITH conversion_stats AS (
                SELECT 
                    COUNT(*) as total_count,
                    SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) as success_count,
                    SUM(CASE WHEN status LIKE 'error%' THEN 1 ELSE 0 END) as error_count
                FROM conversion_history
                WHERE user_id = ?
                AND date(conversion_date) = date('now')
            ),
            feedback_stats AS (
                SELECT 
                    AVG(rating) as avg_rating,
                    COUNT(*) as total_feedback
                FROM feedback
                WHERE user_id = ?
                AND date(created_at) = date('now')
            )
            SELECT 
                c.total_count,
                c.success_count,
                c.error_count,
                f.avg_rating,
                f.total_feedback
            FROM conversion_stats c
            CROSS JOIN feedback_stats f
        ''', (user_id, user_id))
        
        stats = c.fetchone()
        
        return {
            'conversion': {
                'total': stats[0] or 0,
                'success': stats[1] or 0,
                'error': stats[2] or 0
            },
            'feedback': {
                'avg_rating': stats[3] or 0,
                'total': stats[4] or 0
            }
        }
    except Exception as e:
        st.error(f"ユーザー統計の取得中にエラーが発生しました: {str(e)}")
        return {
            'conversion': {'total': 0, 'success': 0, 'error': 0},
            'feedback': {'avg_rating': 0, 'total': 0}
        }
    finally:
        conn.close()

def batch_process_conversions(conversions: list) -> bool:
    """変換履歴を一括処理"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # トランザクション開始
        c.execute('BEGIN TRANSACTION')
        
        try:
            # 一括挿入
            c.executemany('''
                INSERT INTO conversion_history 
                (user_id, document_type, document_date, file_name, status, ip_address, user_agent)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', conversions)
            
            conn.commit()
            return True
        except Exception as e:
            conn.rollback()
            raise e
            
    except Exception as e:
        st.error(f"変換履歴の一括処理中にエラーが発生しました: {str(e)}")
        return False
    finally:
        conn.close()

async def process_pdf_async(uploaded_file, document_type: str, document_date: date = None) -> tuple:
    """PDFを非同期で処理"""
    try:
        # ファイルサイズチェック
        file_size = len(uploaded_file.getvalue()) / (1024 * 1024)  # MBに変換
        if file_size > 200:
            raise ValueError("ファイルサイズが200MBを超えています。")

        # PDFを画像に変換
        pdf_bytes = uploaded_file.getvalue()
        images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1)
        
        if not images:
            raise ValueError("PDFの読み込みに失敗しました。")

        # 画像をバイトデータに変換
        img_byte_arr = io.BytesIO()
        images[0].save(img_byte_arr, format='PNG')
        img_bytes = img_byte_arr.getvalue()

        # まずpdfplumberでテキスト抽出を試みる
        try:
            with pdfplumber.open(uploaded_file) as pdf:
                page = pdf.pages[0]
                text_content = page.extract_text()
                if text_content and len(text_content.strip()) > 0:
                    return await create_excel_file_async(text_content, document_type, document_date)
        except Exception:
            st.warning("テキストの直接抽出に失敗しました。OCR処理を試みます。")

        # OCR処理を非同期で実行
        text_content = await process_pdf_with_ocr_async(img_bytes, document_type)
        if not text_content:
            raise ValueError("PDFからテキストを抽出できませんでした。")

        return await create_excel_file_async(text_content, document_type, document_date)

    except Exception as e:
        st.error(f"PDF処理中にエラーが発生しました: {str(e)}")
        return None, None

async def process_pdf_with_ocr_async(image_bytes: bytes, document_type: str) -> str:
    """OCR処理を非同期で実行"""
    try:
        # 画像をbase64エンコード
        image_base64 = base64.b64encode(image_bytes).decode('utf-8')
        
        # APIキーの取得
        api_key = os.getenv('GOOGLE_VISION_API_KEY')
        if not api_key:
            raise ValueError("Google Cloud Vision APIの設定が見つかりません。")
            
        # APIエンドポイントの設定
        url = f"https://vision.googleapis.com/v1/images:annotate?key={api_key}"
        headers = {"Content-Type": "application/json"}
        
        # リクエストデータの作成
        data = {
            "requests": [{
                "image": {"content": image_base64},
                "features": [{"type": "DOCUMENT_TEXT_DETECTION"}]
            }]
        }
        
        # 非同期でAPIリクエストを送信
        async with aiohttp.ClientSession() as session:
            async with session.post(url, headers=headers, json=data) as response:
                if response.status != 200:
                    raise ValueError("OCR処理に失敗しました。")
                    
                result = await response.json()
                if 'responses' in result and result['responses']:
                    text_annotations = result['responses'][0].get('textAnnotations', [])
                    if text_annotations:
                        return text_annotations[0].get('description', '')
                        
        return None
        
    except Exception as e:
        st.error(f"OCR処理中にエラーが発生しました: {str(e)}")
        return None

async def create_excel_file_async(text_content: str, document_type: str, document_date: date = None) -> tuple:
    """Excelファイルを非同期で作成"""
    try:
        # ドキュメントタイプの日本語名を取得
        doc_type_ja = get_document_type_label(document_type)
        
        # 日付の処理
        if document_date:
            date_str = document_date.strftime("%Y%m%d")
        else:
            date_str = datetime.now().strftime("%Y%m%d")
        
        # Excelワークブックの作成（非同期で実行）
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor() as pool:
            wb = await loop.run_in_executor(pool, create_excel_workbook, text_content, doc_type_ja, date_str)
        
        # メモリ上にExcelファイルを保存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # ファイル名の生成
        file_name = f"{doc_type_ja}_{date_str}.xlsx"
        
        return excel_buffer, file_name
    except Exception as e:
        st.error(f"Excelファイルの作成中にエラーが発生しました: {str(e)}")
        return None, None

def create_excel_workbook(text_content: str, doc_type_ja: str, date_str: str) -> Workbook:
    """Excelワークブックを作成（同期的な処理）"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{doc_type_ja}_{date_str}"
    
    # フォントとスタイルの設定
    default_font = Font(name='Yu Gothic', size=11)
    header_font = Font(name='Yu Gothic', size=12, bold=True)
    title_font = Font(name='Yu Gothic', size=14, bold=True)
    
    # 基本的なセルスタイル
    normal_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 罫線スタイル
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # テキストを行に分割
    lines = text_content.split('\n')
    
    # 行の高さと列幅の初期設定
    ws.row_dimensions[1].height = 30
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # データの書き込みと書式設定
    for i, line in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=line)
        cell = ws[f"A{i}"]
        cell.font = default_font
        cell.alignment = normal_alignment
        cell.border = thin_border
        
        # 金額と思われる部分は右寄せに
        if any(char in line for char in ['¥', '円', '税']):
            cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # タイトル行の特別な書式設定
    if len(lines) > 0:
        title_cell = ws['A1']
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        title_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    return wb

async def process_pdf_batch_async(files: list, document_type: str, document_date: date) -> list:
    """複数のPDFファイルを非同期で一括処理"""
    tasks = []
    for file in files:
        task = process_pdf_async(file, document_type, document_date)
        tasks.append(task)
    
    results = await asyncio.gather(*tasks, return_exceptions=True)
    return [result for result in results if isinstance(result, tuple) and result[0] is not None]

def create_sidebar():
    """サイドバーを作成"""
    with st.sidebar:
        st.markdown("### 👤 アカウント")
        if st.session_state.logged_in:
            st.markdown(f"ようこそ、{st.session_state.username}さん")
            if st.button("ログアウト"):
                handle_logout()
        else:
            # ログインフォーム
            with st.form("login_form"):
                st.markdown("#### ログイン")
                username = st.text_input("ユーザー名")
                password = st.text_input("パスワード", type="password")
                if st.form_submit_button("ログイン"):
                    if handle_login(username, password):
                        st.success("ログインしました！")
                        st.rerun()
            
            # 新規登録フォーム
            with st.form("register_form"):
                st.markdown("#### 新規登録")
                new_username = st.text_input("ユーザー名")
                new_password = st.text_input("パスワード", type="password")
                confirm_password = st.text_input("パスワード（確認）", type="password")
                if st.form_submit_button("登録"):
                    if new_password == confirm_password:
                        if handle_register(new_username, new_password, None):
                            st.success("アカウントが作成されました！")
                            st.rerun()
                    else:
                        st.error("パスワードが一致しません")
        
        st.markdown("### 📄 関連リンク")
        st.markdown("""
        - [利用規約](/terms)
        - [プライバシーポリシー](/privacy)
        - [お問い合わせ](/contact)
        """)

def main():
    """メイン関数"""
    try:
        # データベースの最適化
        optimize_database()
        
        # サイドバーの作成
        create_sidebar()
        
        create_hero_section()
        
        # 変換回数の表示（最上部）
        display_conversion_count()
        
        # ログインユーザーの場合、アラートセクションを表示
        if st.session_state.logged_in:
            create_alert_section()
            check_conversion_limits()
            check_system_status()
            
            # ユーザー統計の表示
            user_stats = get_user_stats(st.session_state.user_id)
            st.sidebar.markdown("### 📊 本日の利用状況")
            
            # 統計情報の計算
            total_conversions = user_stats['conversion']['total']
            success_rate = 0.0
            if total_conversions > 0:
                success_rate = (user_stats['conversion']['success'] / total_conversions * 100)
            
            avg_rating = "なし"
            if user_stats['feedback']['total'] > 0:
                avg_rating = f"{user_stats['feedback']['avg_rating']:.1f}"
            
            # 統計情報の表示
            stats_text = f"""
            - 変換回数: {total_conversions}回
            - 成功率: {success_rate:.1f}%
            - 平均評価: {avg_rating}
            """
            st.sidebar.markdown(stats_text)
        
        # 2カラムレイアウト
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("ファイルをアップロード")
            
            # ドキュメントタイプの選択（ボタン形式）
            document_type = create_document_type_buttons()
            
            # 日付入力
            document_date = st.date_input(
                "書類の日付",
                value=None,
                help="YYYY/MM/DD形式で入力してください"
            )
            
            # ファイルアップロード（複数ファイル対応）
            uploaded_files = st.file_uploader(
                "クリックまたはドラッグ&ドロップでPDFファイルを選択",
                type=['pdf'],
                accept_multiple_files=True,
                help="ファイルサイズの制限: 200MB"
            )
            
            # 広告表示（ファイルアップロードボタンの下）
            show_ads()
            
            st.info("💡 無料プランでは1ページ目のみ変換されます。全ページ変換は有料プランでご利用いただけます。")
            
            if uploaded_files and document_type is not None:
                if st.button("Excelに変換する"):
                    if not st.session_state.processing_pdf:  # 処理中でない場合のみ実行
                        if check_and_increment_conversion_count(st.session_state.get('user_id')):
                            try:
                                # 非同期処理の実行
                                results = asyncio.run(process_pdf_batch_async(uploaded_files, document_type, document_date))
                                
                                # 変換履歴の一括保存
                                if st.session_state.logged_in:
                                    conversions = []
                                    for excel_data, file_name in results:
                                        conversions.append((
                                            st.session_state.user_id,
                                            document_type,
                                            document_date.strftime('%Y-%m-%d') if document_date else None,
                                            file_name,
                                            'success',
                                            st.session_state.get('client_ip', 'unknown'),
                                            st.session_state.get('user_agent', 'unknown')
                                        ))
                                    batch_process_conversions(conversions)
                                
                                # 結果の表示とダウンロード
                                for excel_data, file_name in results:
                                    if excel_data and file_name:
                                        st.download_button(
                                            label=f"{file_name}をダウンロード",
                                            data=excel_data.getvalue(),
                                            file_name=file_name,
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                        )
                                
                                # 変換成功後にフィードバックを促す
                                if st.session_state.logged_in:
                                    st.markdown("---")
                                    create_feedback_section()
                                    
                            except Exception as e:
                                handle_error(e, {
                                    'files': [f.name for f in uploaded_files],
                                    'document_type': document_type,
                                    'document_date': document_date
                                })
        
        with col2:
            st.subheader("プレビュー")
            if uploaded_files:
                preview_image = create_preview(uploaded_files[0])
                if preview_image is not None:
                    st.image(preview_image, use_container_width=True)
        
        # ログインユーザーの場合、レポートセクションを表示
        if st.session_state.logged_in:
            st.markdown("---")
            create_report_section()
        
        create_footer()
        
    except Exception as e:
        handle_error(e, {'context': 'main_function'})

# メイン処理部分
if st.session_state.conversion_success:
    st.session_state.conversion_success = False
    st.session_state.rerun_count = 0

if __name__ == "__main__":
    main() 